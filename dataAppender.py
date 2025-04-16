import openpyxl
from datetime import datetime
import random

# Get the current time
current_time = datetime.now()

# Format the date and time separately
formatted_date = current_time.strftime("%Y-%m-%d")
formatted_time = current_time.strftime("%H:%M:%S")

# Load the existing workbook
file_path = "sensorData.xlsx"

# If the file path doesn't exist, create a new workbook
try:
    wb = openpyxl.load_workbook(file_path)
except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sensor Data"
    ws.append(["Date", "Time", "Value"])
    wb.save(file_path)
    wb = openpyxl.load_workbook(file_path)

ws = wb.active

data = 0 # Set the value of this variable to the data you want to append
#-----DATA TO APPEND-----
data = random.randint(1, 100)
new_data = [
    [formatted_date, formatted_time, data]
]
#-----DATA TO APPEND-----

# Append the new data to the worksheet
ws.append(new_data[0])

# Save the workbook
wb.save(file_path)

print(f"Data appended to {file_path}")