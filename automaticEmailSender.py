import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side

# Email credentials
email_sender = 'dimensionmonitoring839c@gmail.com'
email_password = "riew vgru bvbc urdv"
email_receivers = ['dimensionmonitoring839c@gmail.com', 'recipent@example.com']  # Put your list of email recipients here

# Path to the attachment file
file_path = "sensorData.xlsx"

# Read the spreadsheet file and get the data from the second row, first column and the last row, first column
# Retrieve the data as a string
wb = openpyxl.load_workbook(file_path)
ws = wb.active
start_date = ws.cell(row=2, column=1).value
end_date = ws.cell(row=ws.max_row, column=1).value

#----------------------------------------------
# SPREADSHEET FORMATTTING - ALL DATA SHOULD BE ADDED IN ADVANCE
#----------------------------------------------

#----------------------BORDER FORMATTING----------------------

# Define the border style
column_border_thin = Border(
    right=Side(style='thin')
)

column_border_thick = Border(
    right=Side(style='thick')
)

row_border = Border(
    bottom=Side(style='thick')
)

# Apply the thin border to all columns
for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in col:
        cell.border = column_border_thin

# Apply the thick border to the first column
for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
    for cell in col:
        cell.border = column_border_thick

# Apply the border to the first row
for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = row_border

# Apply the border to the top left cell
ws['B1'].border = Border(
    right=Side(style='thick'),
    bottom=Side(style='thick')
)

# Adjust the column width to fit the content
for col in ws.columns:
    max_length = 0
    column = openpyxl.utils.get_column_letter(col[0].column)
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

#----------------------TEXT FORMATTING----------------------

# Center the text in all cells
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

# Italicize the first and second column
for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
    for cell in col:
        cell.font = openpyxl.styles.Font(italic=True)

# Bold the first row and remove any italic formatting
for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.font = openpyxl.styles.Font(bold=True, italic=False)

wb.save(file_path)

# New filename for the attachment
new_filename = f"{start_date} - {end_date}.xlsx"

def attachFile(file_path, msg):
    # Open the file in binary mode
    with open(file_path, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
        encoders.encode_base64(part)  # Encode the attachment

        # Add header and attach the file
        part.add_header(
            "Content-Disposition",
            f"attachment; filename= {new_filename}",
        )
        msg.attach(part)

def sendEmail(subject, body, email_sender, email_password, email_receivers):
    # Create the email
    msg = MIMEMultipart()
    msg["From"] = email_sender
    msg["To"] = ", ".join(email_receivers)  # Join multiple email addresses with a comma
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    # Attach the file
    attachFile(file_path, msg)

    # Send the email via SMTP server
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(email_sender, email_password)
        server.sendmail(email_sender, email_receivers, msg.as_string())  # Send to all recipients
        server.quit()
        print("Email sent successfully!")
    except Exception as e:
        print(f"Error: {e}")

# Example usage
subject = f"Sensor Data Report From {start_date} to {end_date}"
body = """
The sensor data report for the specified date range can be found in the attached spreadsheet file.
"""
sendEmail(subject, body, email_sender, email_password, email_receivers)

# Clear the entire spreadsheet
ws.delete_rows(2, ws.max_row)
wb.save(file_path)
wb.close()
